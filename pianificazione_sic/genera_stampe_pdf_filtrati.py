#!/usr/bin/env python3
"""
Funzioni per generazione PDF con filtri personalizzati
"""

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
from collections import defaultdict
import calendar
import os

# Importa la funzione turno_a_orario da genera_stampe_pdf
from genera_stampe_pdf import turno_a_orario


def carica_dati_excel(filename='Pianificazione_Corsi_2026.xlsx'):
    """Carica tutti i dati dal file Excel"""
    wb = load_workbook(filename)
    ws = wb['2026']
    
    dati = []
    ultima_data = None  # Memorizza l'ultima data valida per i pomeriggi
    
    for row in range(1, ws.max_row + 1):
        turno = ws.cell(row=row, column=2).value
        
        if turno not in ['mattina', 'Pomeriggio']:
            continue
        
        data = ws.cell(row=row, column=1).value
        
        # Se non c'è data e il turno è Pomeriggio, usa l'ultima data valida
        if not data and turno == 'Pomeriggio' and ultima_data:
            data = ultima_data
        elif not data:
            continue
        
        # Aggiorna l'ultima data valida se presente
        if ws.cell(row=row, column=1).value:
            ultima_data = data
        
        riga_dati = {
            'data': data,
            'turno': turno,
            'percorsi': []
        }
        
        # PERCORSO 1 (C-H)
        perc1 = {
            'nome': ws.cell(row=row, column=3).value,
            'formatore1': ws.cell(row=row, column=4).value,
            'formatore2': ws.cell(row=row, column=5).value,
            'aula': ws.cell(row=row, column=6).value,
            'attivita': ws.cell(row=row, column=7).value,
            'test': ws.cell(row=row, column=8).value,
        }
        if any(perc1.values()):
            riga_dati['percorsi'].append(('Percorso 1', perc1))
        
        # PERCORSO 2 (I-N)
        perc2 = {
            'nome': ws.cell(row=row, column=9).value,
            'formatore1': ws.cell(row=row, column=10).value,
            'formatore2': ws.cell(row=row, column=11).value,
            'aula': ws.cell(row=row, column=12).value,
            'attivita': ws.cell(row=row, column=13).value,
            'test': ws.cell(row=row, column=14).value,
        }
        if any(perc2.values()):
            riga_dati['percorsi'].append(('Percorso 2', perc2))
        
        # PERCORSO 3 (O-T)
        perc3 = {
            'nome': ws.cell(row=row, column=15).value,
            'formatore1': ws.cell(row=row, column=16).value,
            'formatore2': ws.cell(row=row, column=17).value,
            'aula': ws.cell(row=row, column=18).value,
            'attivita': ws.cell(row=row, column=19).value,
            'test': ws.cell(row=row, column=20).value,
        }
        if any(perc3.values()):
            riga_dati['percorsi'].append(('Percorso 3', perc3))
        
        # PERCORSO 4 (U-Z)
        perc4 = {
            'nome': ws.cell(row=row, column=21).value,
            'formatore1': ws.cell(row=row, column=22).value,
            'formatore2': ws.cell(row=row, column=23).value,
            'aula': ws.cell(row=row, column=24).value,
            'attivita': ws.cell(row=row, column=25).value,
            'test': ws.cell(row=row, column=26).value,
        }
        if any(perc4.values()):
            riga_dati['percorsi'].append(('Percorso 4', perc4))
        
        # Fuori aula - con mappatura attività
        fuori_aula_list = []
        for col in range(27, 38, 2):  # 27-37 dispari = formatori, pari = attività
            form = ws.cell(row=row, column=col).value
            att = ws.cell(row=row, column=col+1).value
            if form:
                fuori_aula_list.append({
                    'formatore': form,
                    'attivita': att if att else ''
                })
        
        if fuori_aula_list:
            riga_dati['fuori_aula'] = fuori_aula_list
        
        dati.append(riga_dati)
    
    return dati


def genera_report_aule_settimane(settimane, output_dir='stampe_pdf'):
    """Genera report aule per settimane specifiche"""
    print(f"📋 Generazione report aule per settimane: {settimane}")
    
    dati = carica_dati_excel()
    os.makedirs(output_dir, exist_ok=True)
    
    # Filtra dati per settimane
    dati_filtrati = []
    for riga in dati:
        if riga['data']:
            num_settimana = riga['data'].isocalendar()[1]
            if num_settimana in settimane:
                dati_filtrati.append(riga)
    
    if not dati_filtrati:
        print("⚠️  Nessun dato trovato per le settimane selezionate")
        return
    
    # Organizza per aula
    aule_dati = defaultdict(list)
    
    for riga in dati_filtrati:
        for _, perc in riga['percorsi']:
            aula = perc.get('aula')
            if aula:
                aule_dati[aula].append({
                    'data': riga['data'],
                    'turno': riga['turno'],
                    'percorso': perc.get('nome'),
                    'attivita': perc.get('attivita'),
                    'formatori': [f for f in [perc.get('formatore1'), perc.get('formatore2')] if f]
                })
    
    # Genera PDF
    settimane_str = '_'.join(map(str, sorted(settimane)))
    filename = f"{output_dir}/Prenotazione_Aule_Settimane_{settimane_str}_2026.pdf"
    
    doc = SimpleDocTemplate(filename, pagesize=A4,
                           leftMargin=1.5*cm, rightMargin=1.5*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"<b>PRENOTAZIONE AULE - Settimane {', '.join(map(str, sorted(settimane)))}</b>", title_style))
    elements.append(Paragraph("Anno 2026", styles['Normal']))
    elements.append(Spacer(1, 0.8*cm))
    
    for aula in sorted(aule_dati.keys()):
        aula_style = ParagraphStyle(
            'AulaTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#2E75B5'),
            spaceBefore=20,
            spaceAfter=15
        )
        elements.append(Paragraph(f"<b>AULA {aula}</b>", aula_style))
        
        prenotazioni = sorted(aule_dati[aula], key=lambda x: (x['data'], x['turno']))
        
        table_data = [['Data', 'Turno', 'Percorso', 'Attività', 'Formatori']]
        
        for pren in prenotazioni:
            table_data.append([
                pren['data'].strftime('%d/%m/%Y'),
                pren['turno'],
                pren['percorso'] or '-',
                pren['attivita'] or '-',
                ', '.join(pren['formatori']) if pren['formatori'] else '-'
            ])
        
        table = Table(table_data, colWidths=[2.5*cm, 2.5*cm, 3*cm, 3*cm, 5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.6*cm))
    
    doc.build(elements)
    print(f"✅ PDF generato: {filename}")


def genera_report_formatore_periodo(formatore, data_inizio_str=None, data_fine_str=None, output_dir='stampe_pdf'):
    """Genera report per formatore specifico con periodo"""
    print(f"👤 Generazione report per {formatore}")
    
    dati = carica_dati_excel()
    os.makedirs(output_dir, exist_ok=True)
    
    # Parse date
    data_inizio = datetime.strptime(data_inizio_str, '%Y-%m-%d') if data_inizio_str else None
    data_fine = datetime.strptime(data_fine_str, '%Y-%m-%d') if data_fine_str else None
    
    # Filtra dati
    dati_filtrati = []
    for riga in dati:
        if not riga['data']:
            continue
        
        # Controllo periodo
        if data_inizio and riga['data'] < data_inizio:
            continue
        if data_fine and riga['data'] > data_fine:
            continue
        
        # Controllo formatore
        presente = False
        for _, perc in riga['percorsi']:
            if perc.get('formatore1') == formatore or perc.get('formatore2') == formatore:
                presente = True
                break
        
        if not presente and riga.get('fuori_aula'):
            if formatore in riga['fuori_aula']:
                presente = True
        
        if presente:
            dati_filtrati.append(riga)
    
    if not dati_filtrati:
        print(f"⚠️  Nessun turno trovato per {formatore}")
        return
    
    # Genera PDF
    periodo_str = ""
    if data_inizio and data_fine:
        periodo_str = f"_{data_inizio.strftime('%d%m')}-{data_fine.strftime('%d%m')}"
    elif data_inizio:
        periodo_str = f"_dal_{data_inizio.strftime('%d%m')}"
    elif data_fine:
        periodo_str = f"_al_{data_fine.strftime('%d%m')}"
    
    filename = f"{output_dir}/Programma_Formatore_{formatore}{periodo_str}_2026.pdf"
    
    doc = SimpleDocTemplate(filename, pagesize=A4,
                           leftMargin=2*cm, rightMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    periodo_text = "Anno 2026"
    if data_inizio and data_fine:
        periodo_text = f"Dal {data_inizio.strftime('%d/%m/%Y')} al {data_fine.strftime('%d/%m/%Y')}"
    elif data_inizio:
        periodo_text = f"Dal {data_inizio.strftime('%d/%m/%Y')}"
    elif data_fine:
        periodo_text = f"Al {data_fine.strftime('%d/%m/%Y')}"
    
    elements.append(Paragraph(f"<b>PROGRAMMA FORMATORE {formatore}</b>", title_style))
    elements.append(Paragraph(periodo_text, styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    table_data = [['Data', 'Turno', 'Percorso', 'Aula', 'Attività', 'Note']]
    
    for riga in dati_filtrati:
        data_str = riga['data'].strftime('%d/%m/%Y')
        turno = riga['turno']
        
        for _, perc in riga['percorsi']:
            if perc.get('formatore1') == formatore or perc.get('formatore2') == formatore:
                ruolo = "Form1" if perc.get('formatore1') == formatore else "Form2"
                altro_form = perc.get('formatore2') if ruolo == "Form1" else perc.get('formatore1')
                note = f"Con: {altro_form}" if altro_form else ""
                
                table_data.append([
                    data_str,
                    turno,
                    perc.get('nome', '-'),
                    perc.get('aula', '-'),
                    perc.get('attivita', '-'),
                    note
                ])
        
        if riga.get('fuori_aula') and formatore in riga['fuori_aula']:
            idx = riga['fuori_aula'].index(formatore)
            att = riga.get('attivita_esterne', [])[idx] if idx < len(riga.get('attivita_esterne', [])) else '-'
            table_data.append([
                data_str,
                turno,
                'FUORI AULA',
                '-',
                att,
                ''
            ])
    
    table = Table(table_data, colWidths=[3*cm, 2.5*cm, 3.5*cm, 2.5*cm, 4*cm, 4.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    print(f"✅ PDF generato: {filename}")
    print(f"   Turni: {len(dati_filtrati)}")


def genera_report_corso_specifico(corso, output_dir='stampe_pdf'):
    """Genera report per corso specifico"""
    print(f"📖 Generazione report per corso: {corso}")
    
    dati = carica_dati_excel()
    os.makedirs(output_dir, exist_ok=True)
    
    # Filtra dati per corso
    dati_filtrati = []
    for riga in dati:
        if not riga['data']:
            continue
        
        for _, perc in riga['percorsi']:
            if perc.get('nome') == corso:
                dati_filtrati.append((riga, perc))
                break
    
    if not dati_filtrati:
        print(f"⚠️  Nessun turno trovato per {corso}")
        return
    
    # Genera PDF
    filename = f"{output_dir}/Programma_Corso_{corso.replace('/', '-')}_2026.pdf"
    
    doc = SimpleDocTemplate(filename, pagesize=A4,
                           leftMargin=2*cm, rightMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    prima_data = dati_filtrati[0][0]['data']
    ultima_data = dati_filtrati[-1][0]['data']
    
    elements.append(Paragraph(f"<b>PROGRAMMA CORSO: {corso}</b>", title_style))
    elements.append(Paragraph(f"Dal {prima_data.strftime('%d/%m/%Y')} al {ultima_data.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    # Organizza turni per giorno (mattina e pomeriggio separati)
    giorni_dict = defaultdict(lambda: {'mattina': None, 'pomeriggio': None})
    
    for riga, perc in dati_filtrati:
        data_key = riga['data'].date()
        turno_tipo = riga['turno'].lower()
        giorni_dict[data_key][turno_tipo] = (riga, perc)
    
    # Tabella: colonne per ogni giorno, righe per mattina/pomeriggio
    # MASSIMO 5 GIORNI PER TABELLA
    giorni_ordinati = sorted(giorni_dict.keys())
    
    # Dividi i giorni in gruppi di massimo 5
    for gruppo_idx in range(0, len(giorni_ordinati), 5):
        giorni_gruppo = giorni_ordinati[gruppo_idx:gruppo_idx + 5]
        
        if gruppo_idx > 0:
            # Aggiungi spazio tra le tabelle
            elements.append(Spacer(1, 0.8*cm))
        
        # Header: Date
        header_row = ['Orario'] + [g.strftime('%d/%m\n%a')[:10] for g in giorni_gruppo]
        table_data = [header_row]
        
        # Riga Mattina
        mattina_row = ['Mattina\n' + turno_a_orario('mattina')]
        for giorno in giorni_gruppo:
            data = giorni_dict[giorno]['mattina']
            if data:
                riga, perc = data
                formatori = []
                if perc.get('formatore1'):
                    formatori.append(perc['formatore1'])
                if perc.get('formatore2'):
                    formatori.append(perc['formatore2'])
                
                # Formato migliorato con etichette
                attivita = perc.get('attivita', '-')
                aula = perc.get('aula', '-')
                
                cell_text = f"<b>Attività:</b> {attivita}\n"
                cell_text += f"<b>Aula:</b> {aula}\n"
                if formatori:
                    cell_text += f"<b>Formatore:</b>\n" + '\n'.join(formatori)
                else:
                    cell_text += "<b>Formatore:</b> -"
                
                # Usa Paragraph per supportare tag HTML
                cell_style = ParagraphStyle(
                    'CellStyle',
                    parent=styles['Normal'],
                    fontSize=7,
                    alignment=TA_CENTER,
                    leading=9
                )
                mattina_row.append(Paragraph(cell_text, cell_style))
            else:
                mattina_row.append('-')
        
        table_data.append(mattina_row)
        
        # Riga Pomeriggio
        pomeriggio_row = ['Pomeriggio\n' + turno_a_orario('Pomeriggio')]
        for giorno in giorni_gruppo:
            data = giorni_dict[giorno]['pomeriggio']
            if data:
                riga, perc = data
                formatori = []
                if perc.get('formatore1'):
                    formatori.append(perc['formatore1'])
                if perc.get('formatore2'):
                    formatori.append(perc['formatore2'])
                
                # Formato migliorato con etichette
                attivita = perc.get('attivita', '-')
                aula = perc.get('aula', '-')
                
                cell_text = f"<b>Attività:</b> {attivita}\n"
                cell_text += f"<b>Aula:</b> {aula}\n"
                if formatori:
                    cell_text += f"<b>Formatore:</b>\n" + '\n'.join(formatori)
                else:
                    cell_text += "<b>Formatore:</b> -"
                
                # Usa Paragraph per supportare tag HTML
                cell_style = ParagraphStyle(
                    'CellStyle',
                    parent=styles['Normal'],
                    fontSize=7,
                    alignment=TA_CENTER,
                    leading=9
                )
                pomeriggio_row.append(Paragraph(cell_text, cell_style))
            else:
                pomeriggio_row.append('-')
        
        table_data.append(pomeriggio_row)
        
        # Calcola larghezza colonne: massimo 5 giorni
        num_giorni = len(giorni_gruppo)
        col_width = 3*cm  # Larghezza fissa per 5 giorni
        
        col_widths = [2.5*cm] + [col_width] * num_giorni
        
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D9E1F2')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 1), (-1, -1), 3),
            ('RIGHTPADDING', (0, 1), (-1, -1), 3),
        ]))
        
        elements.append(table)
    
    # Legenda attività (più piccola)
    elements.append(Spacer(1, 0.8*cm))
    
    legend_title_style = ParagraphStyle(
        'LegendTitle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#1F4E78'),
    )
    
    elements.append(Paragraph("<b>LEGENDA ATTIVITÀ:</b>", legend_title_style))
    elements.append(Spacer(1, 0.1*cm))
    
    # Legenda su 2 colonne
    legenda_data = [
        ['AULA', 'Lezione in aula', 'TT', 'Test Tedesco'],
        ['DIGI', 'Attività digitali', 'TI', 'Test Inglese'],
        ['CV', 'Curriculum Vitae', 'UFF', 'Ufficio'],
        ['CS', 'Colloquio selezione', 'COL', 'Colloquio'],
        ['RA', 'Ricerca attiva', 'C', 'Corso']
    ]
    
    legenda_table = Table(legenda_data, colWidths=[1*cm, 3.5*cm, 1*cm, 3.5*cm])
    legenda_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F0F8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E8F0F8')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(legenda_table)
    
    doc.build(elements)
    
    print(f"✅ PDF generato: {filename}")
