#!/usr/bin/env python3
"""
GENERATORE PDF INTERATTIVO - Sistema con parametri per interfaccia web
"""

import sys
import json
import re
from openpyxl import load_workbook
from genera_stampe_pdf import (
    carica_dati_excel,
    genera_report_aule,
    genera_report_formatori,
    genera_report_corsi,
    genera_report_settimanale
)
from genera_stampe_pdf_filtrati import (
    genera_report_aule_settimane,
    genera_report_formatore_periodo,
    genera_report_corso_specifico
)


def get_lista_corsi():
    """Restituisce lista corsi disponibili dalle colonne percorso (C, I, O, U)"""
    try:
        wb = load_workbook('Pianificazione_Corsi_2026.xlsx')
        ws = wb['2026']
        
        percorsi = set()
        # Colonne percorso: C(3), I(9), O(15), U(21)
        for row in range(6, ws.max_row + 1):
            for col in [3, 9, 15, 21]:
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip():
                    val_str = str(val).strip()
                    # Esclude intestazioni e valori non validi
                    if val_str.lower() not in ['percorso', 'bcc', 'none']:
                        percorsi.add(val_str)
        
        # Ordina naturalmente (1a, 1b, 2a, 2b, ..., 10a, 10b, ...)
        def sort_key(x):
            match = re.match(r'(\d+)([a-z])', x.lower())
            if match:
                return (int(match.group(1)), match.group(2))
            return (999, x)
        
        return sorted(percorsi, key=sort_key)
    except Exception as e:
        print(f"Errore caricamento corsi: {e}", file=sys.stderr)
        return []


def get_lista_formatori():
    """Restituisce lista formatori disponibili"""
    # Usa la lista formatori da genera_stampe_pdf.py
    from genera_stampe_pdf import FORMATORI
    return sorted(FORMATORI)


def get_lista_settimane():
    """Restituisce lista settimane disponibili"""
    try:
        dati = carica_dati_excel('Pianificazione_Corsi_2026.xlsx')
        settimane = set()
        for riga in dati:
            if riga['data']:
                num_settimana = riga['data'].isocalendar()[1]
                settimane.add(num_settimana)
        return sorted(settimane)
    except Exception as e:
        return []


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Comando non specificato'}))
        sys.exit(1)
    
    comando = sys.argv[1]
    
    if comando == 'lista_corsi':
        corsi = get_lista_corsi()
        print(json.dumps({'corsi': corsi}))
    
    elif comando == 'lista_formatori':
        formatori = get_lista_formatori()
        print(json.dumps({'formatori': formatori}))
    
    elif comando == 'lista_settimane':
        settimane = get_lista_settimane()
        print(json.dumps({'settimane': settimane}))
    
    elif comando == 'genera_aule_settimane':
        # Parametro: settimane (es: "1,2,3" o "1-5")
        if len(sys.argv) < 3:
            print(json.dumps({'error': 'Settimane non specificate'}))
            sys.exit(1)
        
        settimane_str = sys.argv[2]
        settimane = [int(s.strip()) for s in settimane_str.split(',')]
        
        try:
            genera_report_aule_settimane(settimane)
            print(json.dumps({'success': True, 'message': f'PDF generato per settimane {settimane}'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    elif comando == 'genera_aule_periodo':
        # Parametri: data_inizio, data_fine (formato YYYY-MM-DD)
        if len(sys.argv) < 4:
            print(json.dumps({'error': 'Date non specificate'}))
            sys.exit(1)
        
        from datetime import datetime
        data_inizio_str = sys.argv[2]
        data_fine_str = sys.argv[3]
        
        try:
            # Converti da YYYY-MM-DD a datetime
            data_inizio = datetime.strptime(data_inizio_str, '%Y-%m-%d')
            data_fine = datetime.strptime(data_fine_str, '%Y-%m-%d')
            
            # Carica dati e genera PDF
            dati = carica_dati_excel('Pianificazione_Corsi_2026.xlsx')
            genera_report_aule(dati, 'stampe_pdf', data_inizio, data_fine)
            
            print(json.dumps({'success': True, 'message': f'PDF generato per periodo {data_inizio_str} - {data_fine_str}'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    elif comando == 'genera_formatore':
        # Parametri: formatore, data_inizio, data_fine
        if len(sys.argv) < 4:
            print(json.dumps({'error': 'Parametri mancanti'}))
            sys.exit(1)
        
        formatore = sys.argv[2]
        data_inizio = sys.argv[3] if sys.argv[3] != 'null' else None
        data_fine = sys.argv[4] if len(sys.argv) > 4 and sys.argv[4] != 'null' else None
        
        try:
            genera_report_formatore_periodo(formatore, data_inizio, data_fine)
            print(json.dumps({'success': True, 'message': f'PDF generato per {formatore}'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    elif comando == 'genera_corso':
        # Parametro: nome corso
        if len(sys.argv) < 3:
            print(json.dumps({'error': 'Corso non specificato'}))
            sys.exit(1)
        
        corso = sys.argv[2]
        
        try:
            genera_report_corso_specifico(corso)
            print(json.dumps({'success': True, 'message': f'PDF generato per {corso}'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    elif comando == 'genera_tutti_formatori':
        try:
            dati = carica_dati_excel('Pianificazione_Corsi_2026.xlsx')
            genera_report_formatori(dati)
            print(json.dumps({'success': True, 'message': 'PDF generati per tutti i formatori'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    elif comando == 'genera_settimanale':
        try:
            dati = carica_dati_excel('Pianificazione_Corsi_2026.xlsx')
            genera_report_settimanale(dati)
            print(json.dumps({'success': True, 'message': 'PDF settimanali generati'}))
        except Exception as e:
            print(json.dumps({'error': str(e)}))
    
    else:
        print(json.dumps({'error': f'Comando sconosciuto: {comando}'}))
