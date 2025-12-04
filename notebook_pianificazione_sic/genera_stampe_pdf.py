#!/usr/bin/env python3
"""
GENERATORE STAMPE PDF - PIANIFICAZIONE CORSI 2026
==================================================

Genera 3 tipi di report PDF:
1. Prenotazione Aule (mensile) - per ogni aula
2. Programma Formatori (mensile) - per ogni formatore
3. Programma Corso (5 giorni) - per studenti

"""

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
from collections import defaultdict
import calendar
import os

# Carica configurazione orari dall'Excel
def carica_orari_da_excel(filename='Pianificazione_Corsi_2026.xlsx'):
    """Legge gli orari dal foglio Assumptions dell'Excel"""
    try:
        wb = load_workbook(filename)
        if 'Assumptions' not in wb.sheetnames:
            raise Exception("Foglio Assumptions non trovato")
        
        ws = wb['Assumptions']
        
        # Legge orari dalla riga 4 (Mattina) e riga 5 (Pomeriggio)
        orari = {
            'mattina': {
                'inizio': ws['B4'].value if ws['B4'].value else '09:00',
                'fine': ws['C4'].value if ws['C4'].value else '13:00'
            },
            'pomeriggio': {
                'inizio': ws['B5'].value if ws['B5'].value else '14:00',
                'fine': ws['C5'].value if ws['C5'].value else '18:00'
            }
        }
        return orari
    except Exception as e:
        # Valori di default se c'è un errore
        print(f"⚠️  Impossibile caricare orari da Excel, uso valori di default: {e}")
        return {
            'mattina': {'inizio': '09:00', 'fine': '13:00'},
            'pomeriggio': {'inizio': '14:00', 'fine': '18:00'}
        }

# COSTANTI
FORMATORI = ['CL', 'MC', 'LD', 'EP', 'IP', 'FB', 'GZ', 'DC']
AULE = ['103', '103a', '108', '110', 'UFF']

# Dati formatori per calcolo percentuale
FORMATORI_DATA = {
    'CL': {'perc': 0.7, 'giorni_previsti': 155},
    'MC': {'perc': 0.5, 'giorni_previsti': 111},
    'LD': {'perc': 0.8, 'giorni_previsti': 177},
    'EP': {'perc': 0.9, 'giorni_previsti': 199},
    'IP': {'perc': 0.9, 'giorni_previsti': 199},
    'FB': {'perc': 0.8, 'giorni_previsti': 177},
    'GZ': {'perc': 0.8, 'giorni_previsti': 177},
    'DC': {'perc': 0.4, 'giorni_previsti': 88},
}


def turno_a_orario(turno):
    """Converte 'mattina' o 'Pomeriggio' in formato orario es: '09:00-13:00'"""
    # Ricarica gli orari ogni volta per avere i valori aggiornati dall'Excel
    orari_config = carica_orari_da_excel()
    
    turno_lower = turno.lower() if turno else ''
    if turno_lower == 'mattina':
        return f"{orari_config['mattina']['inizio']}-{orari_config['mattina']['fine']}"
    elif turno_lower == 'pomeriggio':
        return f"{orari_config['pomeriggio']['inizio']}-{orari_config['pomeriggio']['fine']}"
    return turno  # Ritorna il valore originale se non è riconosciuto


def carica_attivita_esterne(filename='Pianificazione_Corsi_2026.xlsx'):
    """Carica le mappature delle attività esterne dal foglio ATT. ESTERNE"""
    try:
        wb = load_workbook(filename)
        ws = wb['ATT. ESTERNE']
        
        mappature = {}
        # Leggi dalla riga 4 in poi (dopo l'header)
        for row in range(4, 20):  # Fino a riga 20
            codice = ws.cell(row=row, column=1).value  # Colonna A
            descrizione = ws.cell(row=row, column=2).value  # Colonna B
            
            if codice and descrizione:
                mappature[str(codice).strip()] = str(descrizione).strip()
        
        wb.close()
        return mappature
    except Exception as e:
        print(f"⚠️  Errore caricamento attività esterne: {e}")
        # Mappature di default
        return {
            'Amm': 'Amministrazione',
            'IA': 'Intelligenza Artificiale',
            'AI': 'Corso AI',
            'Dig': 'Digitale',
            'P1': 'Progetto 1',
            'P2': 'Progetto 2',
            'P3': 'Progetto 3',
            'P4': 'Progetto 4',
            'P5': 'Progetto 5',
        }


def carica_dati_excel(filename):
    """Carica tutti i dati dal file Excel"""
    print(f"📖 Caricamento dati da {filename}...")
    
    wb = load_workbook(filename)
    ws = wb['2026']
    
    dati = []
    ultima_data = None  # Memorizza l'ultima data valida per i pomeriggi
    
    for row in range(1, ws.max_row + 1):
        turno = ws.cell(row=row, column=2).value
        
        if turno not in ['mattina', 'Pomeriggio']:
            continue
        
        data = ws.cell(row=row, column=1).value
        
        # Se non c'è data e il turno è Pomeriggio, usa l'ultima data valida
        if not data and turno == 'Pomeriggio' and ultima_data:
            data = ultima_data
        elif not data:
            continue
        
        # Converti stringa in datetime se necessario
        if isinstance(data, str):
            try:
                data = datetime.strptime(data, '%d/%m/%Y')
            except ValueError:
                continue
        
        # Aggiorna l'ultima data valida se presente
        if ws.cell(row=row, column=1).value:
            ultima_data = data
        
        # Estrai tutti i dati della riga
        riga_dati = {
            'data': data,
            'turno': turno,
            'percorsi': []
        }
        
        # PERCORSO 1 (C-H)
        perc1 = {
            'nome': ws.cell(row=row, column=3).value,
            'formatore1': ws.cell(row=row, column=4).value,
            'formatore2': ws.cell(row=row, column=5).value,
            'aula': ws.cell(row=row, column=6).value,
            'attivita': ws.cell(row=row, column=7).value,
            'test': ws.cell(row=row, column=8).value,
        }
        if any(perc1.values()):
            riga_dati['percorsi'].append(('Percorso 1', perc1))
        
        # PERCORSO 2 (I-N)
        perc2 = {
            'nome': ws.cell(row=row, column=9).value,
            'formatore1': ws.cell(row=row, column=10).value,
            'formatore2': ws.cell(row=row, column=11).value,
            'aula': ws.cell(row=row, column=12).value,
            'attivita': ws.cell(row=row, column=13).value,
            'test': ws.cell(row=row, column=14).value,
        }
        if any(perc2.values()):
            riga_dati['percorsi'].append(('Percorso 2', perc2))
        
        # PERCORSO 3 (O-T)
        perc3 = {
            'nome': ws.cell(row=row, column=15).value,
            'formatore1': ws.cell(row=row, column=16).value,
            'formatore2': ws.cell(row=row, column=17).value,
            'aula': ws.cell(row=row, column=18).value,
            'attivita': ws.cell(row=row, column=19).value,
            'test': ws.cell(row=row, column=20).value,
        }
        if any(perc3.values()):
            riga_dati['percorsi'].append(('Percorso 3', perc3))
        
        # PERCORSO 4 (U-Z)
        perc4 = {
            'nome': ws.cell(row=row, column=21).value,
            'formatore1': ws.cell(row=row, column=22).value,
            'formatore2': ws.cell(row=row, column=23).value,
            'aula': ws.cell(row=row, column=24).value,
            'attivita': ws.cell(row=row, column=25).value,
            'test': ws.cell(row=row, column=26).value,
        }
        if any(perc4.values()):
            riga_dati['percorsi'].append(('Percorso 4', perc4))
        
        # FUORI AULA (AB-AK) - 5 coppie formatore/attività
        formatori_fa = []
        attivita_fa = []
        for i in range(5):
            form_col = 28 + (i * 2)  # 28,30,32,34,36
            att_col = 29 + (i * 2)   # 29,31,33,35,37
            formatore = ws.cell(row=row, column=form_col).value
            attivita = ws.cell(row=row, column=att_col).value
            if formatore:
                formatori_fa.append({'formatore': formatore, 'attivita': attivita or ''})
            if attivita:
                attivita_fa.append(attivita)
        
        riga_dati['fuori_aula'] = formatori_fa
        riga_dati['attivita_esterne'] = attivita_fa
        
        dati.append(riga_dati)
    
    print(f"✅ Caricati {len(dati)} turni\n")
    return dati


def genera_report_aule(dati, output_dir='stampe_pdf', data_inizio=None, data_fine=None):
    """
    REPORT 1: Prenotazione Aule (UN SOLO FILE con TUTTE le aule)
    Opzionalmente filtrabile per periodo
    """
    print("=" * 70)
    if data_inizio and data_fine:
        print(f"📋 REPORT 1: PRENOTAZIONE AULE ({data_inizio.strftime('%d/%m/%Y')} - {data_fine.strftime('%d/%m/%Y')})")
    else:
        print("📋 REPORT 1: PRENOTAZIONE AULE (TUTTE LE AULE)")
    print("=" * 70)
    print()
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Organizza dati per aula e mese
    aule_mensili = defaultdict(lambda: defaultdict(list))
    
    for riga in dati:
        if not riga['data']:
            continue
        
        data = riga['data']
        
        # Filtra per periodo se specificato
        if data_inizio and data < data_inizio:
            continue
        if data_fine and data > data_fine:
            continue
        
        mese = data.month
        
        for perc_label, perc_dati in riga['percorsi']:
            aula = perc_dati.get('aula')
            if not aula:
                continue
            
            aule_mensili[aula][mese].append({
                'data': data,
                'turno': riga['turno'],
                'percorso': perc_dati.get('nome', ''),
                'attivita': perc_dati.get('attivita', ''),
                'formatori': [f for f in [perc_dati.get('formatore1'), perc_dati.get('formatore2')] if f]
            })
    
    # Genera UN SOLO PDF con TUTTE le aule
    if data_inizio and data_fine:
        filename = os.path.join(output_dir, f'Prenotazione_Aule_{data_inizio.strftime("%d%m%Y")}_{data_fine.strftime("%d%m%Y")}.pdf')
    else:
        filename = os.path.join(output_dir, 'Prenotazione_Aule_2026.pdf')
    
    doc = SimpleDocTemplate(filename, pagesize=A4, 
                           leftMargin=1.5*cm, rightMargin=1.5*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Stile titolo principale
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Titolo documento
    if data_inizio and data_fine:
        elements.append(Paragraph(f"<b>PRENOTAZIONE AULE 2026</b>", title_style))
        elements.append(Paragraph(f"Dal {data_inizio.strftime('%d/%m/%Y')} al {data_fine.strftime('%d/%m/%Y')}", styles['Normal']))
    else:
        elements.append(Paragraph(f"<b>PRENOTAZIONE AULE 2026</b>", title_style))
        elements.append(Paragraph(f"Tutte le aule - Anno 2026", styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    # Per ogni aula
    for idx_aula, aula in enumerate(sorted(aule_mensili.keys())):
        # Titolo aula
        aula_style = ParagraphStyle(
            'AulaTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#2E75B5'),
            spaceBefore=20,
            spaceAfter=15,
            alignment=TA_LEFT
        )
        elements.append(Paragraph(f"<b>AULA {aula}</b>", aula_style))
        
        # Per ogni mese
        for mese in sorted(aule_mensili[aula].keys()):
            prenotazioni = aule_mensili[aula][mese]
            
            if not prenotazioni:
                continue
            
            # Titolo mese
            nome_mese = calendar.month_name[mese].upper()
            month_style = ParagraphStyle(
                'MonthTitle',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#2E75B5'),
                spaceBefore=10,
                spaceAfter=10
            )
            elements.append(Paragraph(f"<b>{nome_mese} 2026</b>", month_style))
            
            # Statistiche
            num_prenotazioni = len(prenotazioni)
            num_mattine = sum(1 for p in prenotazioni if p['turno'] == 'mattina')
            num_pomeriggi = sum(1 for p in prenotazioni if p['turno'] == 'Pomeriggio')
            
            attivita_count = defaultdict(int)
            for p in prenotazioni:
                if p['attivita']:
                    attivita_count[p['attivita']] += 1
            
            stats_text = f"<b>Prenotazioni:</b> {num_prenotazioni} (Mattina: {num_mattine}, Pomeriggio: {num_pomeriggi})"
            elements.append(Paragraph(stats_text, styles['Normal']))
            
            if attivita_count:
                att_text = "<b>Attività:</b> " + ", ".join([f"{att} ({cnt})" for att, cnt in sorted(attivita_count.items())])
                elements.append(Paragraph(att_text, styles['Normal']))
            
            elements.append(Spacer(1, 0.3*cm))
            
            # Tabella prenotazioni
            table_data = [['Data', 'Orario', 'Percorso', 'Attività', 'Formatori']]
            
            for pren in sorted(prenotazioni, key=lambda x: (x['data'], x['turno'])):
                table_data.append([
                    pren['data'].strftime('%d/%m/%Y'),
                    turno_a_orario(pren['turno']),
                    pren['percorso'] or '-',
                    pren['attivita'] or '-',
                    ', '.join(pren['formatori']) if pren['formatori'] else '-'
                ])
            
            table = Table(table_data, colWidths=[2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 4.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.6*cm))
        
        # Separatore tra aule (tranne l'ultima)
        if idx_aula < len(aule_mensili) - 1:
            elements.append(PageBreak())
    
    doc.build(elements)
    
    total_prenotazioni = sum(len(aule_mensili[a][m]) for a in aule_mensili for m in aule_mensili[a])
    print(f"   ✅ Generato: {len(aule_mensili)} aule, {total_prenotazioni} prenotazioni totali\n")
    print(f"✅ Report aule completato: {filename}\n")


def genera_report_formatori(dati, output_dir='stampe_pdf'):
    """
    REPORT 2: Programma Formatori (mensile)
    Per ogni formatore: calendario mensile + conteggio ore, rimanenti, percentuale
    """
    print("=" * 70)
    print("👥 REPORT 2: PROGRAMMA FORMATORI (MENSILE)")
    print("=" * 70)
    print()
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Organizza dati per formatore e mese
    formatori_mensili = defaultdict(lambda: defaultdict(list))
    formatori_totali = defaultdict(float)  # Conta turni (0.5 giorni per turno)
    
    for riga in dati:
        if not riga['data']:
            continue
        
        data = riga['data']
        mese = data.month
        
        # Percorsi
        for perc_label, perc_dati in riga['percorsi']:
            for formatore in [perc_dati.get('formatore1'), perc_dati.get('formatore2')]:
                if formatore and formatore in FORMATORI:
                    formatori_mensili[formatore][mese].append({
                        'data': data,
                        'turno': riga['turno'],
                        'percorso': perc_dati.get('nome', ''),
                        'aula': perc_dati.get('aula', ''),
                        'attivita': perc_dati.get('attivita', ''),
                        'tipo': 'Corso'
                    })
                    formatori_totali[formatore] += 0.5
        
        # Fuori aula - con mappatura attività
        for fa_item in riga.get('fuori_aula', []):
            formatore = fa_item.get('formatore') if isinstance(fa_item, dict) else fa_item
            attivita_codice = fa_item.get('attivita', '') if isinstance(fa_item, dict) else ''
            
            if formatore in FORMATORI:
                formatori_mensili[formatore][mese].append({
                    'data': data,
                    'turno': riga['turno'],
                    'percorso': '-',
                    'aula': 'Fuori aula',
                    'attivita': attivita_codice or 'Attività esterna',
                    'tipo': 'Fuori aula'
                })
                formatori_totali[formatore] += 0.5
    
    # Carica mappature attività esterne
    mappature_attivita = carica_attivita_esterne()
    
    # Genera un PDF per ogni formatore
    for formatore in sorted(formatori_mensili.keys()):
        filename = os.path.join(output_dir, f'Programma_Formatore_{formatore}_2026.pdf')
        
        print(f"📄 Generazione: {filename}")
        
        doc = SimpleDocTemplate(filename, pagesize=A4,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titolo
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph(f"<b>PROGRAMMA FORMATORE {formatore}</b>", title_style))
        elements.append(Paragraph(f"Anno 2026", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # STATISTICHE ANNUALI
        giorni_svolti = formatori_totali[formatore]
        turni_svolti = int(giorni_svolti * 2)
        
        if formatore in FORMATORI_DATA:
            giorni_previsti = FORMATORI_DATA[formatore]['giorni_previsti']
            percentuale = FORMATORI_DATA[formatore]['perc']
            giorni_rimanenti = giorni_previsti - giorni_svolti
            perc_svolti = (giorni_svolti / giorni_previsti * 100) if giorni_previsti > 0 else 0
            
            stats_data = [
                ['STATISTICHE ANNUALI', '', '', ''],
                ['% Contratto', 'Giorni Previsti', 'Giorni Svolti', 'Giorni Rimanenti'],
                [f'{int(percentuale*100)}%', f'{giorni_previsti}', f'{giorni_svolti:.1f}', f'{giorni_rimanenti:.1f}'],
                ['Turni Svolti', 'Percentuale Completata', '', ''],
                [f'{turni_svolti}', f'{perc_svolti:.1f}%', '', '']
            ]
            
            stats_table = Table(stats_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D9E1F2')),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#D9E1F2')),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('SPAN', (0, 0), (-1, 0)),
                ('SPAN', (1, 4), (-1, 4)),
            ]))
            
            elements.append(stats_table)
            elements.append(Spacer(1, 1*cm))
        
        # Per ogni mese
        for mese in sorted(formatori_mensili[formatore].keys()):
            impegni = formatori_mensili[formatore][mese]
            
            if not impegni:
                continue
            
            # Titolo mese
            nome_mese = calendar.month_name[mese].upper()
            month_style = ParagraphStyle(
                'MonthTitle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#2E75B5'),
                spaceBefore=10,
                spaceAfter=10
            )
            elements.append(Paragraph(f"<b>{nome_mese} 2026</b>", month_style))
            
            # Conteggio mensile
            giorni_mese = len(impegni) * 0.5
            turni_mese = len(impegni)
            elements.append(Paragraph(f"<b>Impegni:</b> {turni_mese} turni ({giorni_mese:.1f} giorni)", styles['Normal']))
            elements.append(Spacer(1, 0.3*cm))
            
            # Tabella impegni
            table_data = [['Data', 'Orario', 'Percorso', 'Aula', 'Attività']]
            
            for imp in sorted(impegni, key=lambda x: (x['data'], x['turno'])):
                # Mappatura nome attività se è fuori aula
                attivita_display = imp['attivita'] or '-'
                if imp['tipo'] == 'Fuori aula' and imp['attivita'] in mappature_attivita:
                    attivita_display = mappature_attivita[imp['attivita']]
                
                table_data.append([
                    imp['data'].strftime('%d/%m/%Y'),
                    turno_a_orario(imp['turno']),
                    imp['percorso'] or '-',
                    imp['aula'] or '-',
                    attivita_display
                ])
            
            table = Table(table_data, colWidths=[2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 4.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.8*cm))
        
        doc.build(elements)
        print(f"   ✅ Generato per {formatore}: {giorni_svolti:.1f} giorni, {len(formatori_mensili[formatore])} mesi\n")
    
    print(f"✅ Report formatori completati in: {output_dir}/\n")


def genera_report_corsi(dati, output_dir='stampe_pdf'):
    """
    REPORT 3: Programma Corso (5 giorni)
    Per studenti: orari inizio/fine, formatori, aule, test
    """
    print("=" * 70)
    print("📚 REPORT 3: PROGRAMMA CORSI (5 GIORNI)")
    print("=" * 70)
    print()
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Organizza per percorso
    percorsi = defaultdict(list)
    
    for riga in dati:
        if not riga['data']:
            continue
        
        for perc_label, perc_dati in riga['percorsi']:
            nome_percorso = perc_dati.get('nome')
            if not nome_percorso:
                continue
            
            percorsi[nome_percorso].append({
                'data': riga['data'],
                'turno': riga['turno'],
                'formatore1': perc_dati.get('formatore1'),
                'formatore2': perc_dati.get('formatore2'),
                'aula': perc_dati.get('aula'),
                'attivita': perc_dati.get('attivita'),
                'test': perc_dati.get('test')
            })
    
    # Genera PDF per ogni percorso (solo quelli con almeno 3 giorni)
    for nome_percorso in sorted(percorsi.keys()):
        turni = percorsi[nome_percorso]
        
        # Conta giorni unici
        giorni_unici = len(set(t['data'].date() for t in turni))
        
        if giorni_unici < 3:  # Skip percorsi incompleti
            continue
        
        filename = os.path.join(output_dir, f'Programma_Corso_{nome_percorso}_2026.pdf')
        
        print(f"📄 Generazione: {filename}")
        
        doc = SimpleDocTemplate(filename, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2.5*cm, bottomMargin=2.5*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titolo principale
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=15,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph(f"<b>PROGRAMMA CORSO {nome_percorso}</b>", title_style))
        elements.append(Paragraph(f"Anno 2026", styles['Normal']))
        
        # Date corso
        turni_ordinati = sorted(turni, key=lambda x: (x['data'], x['turno']))
        data_inizio = turni_ordinati[0]['data']
        data_fine = turni_ordinati[-1]['data']
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        elements.append(Paragraph(
            f"<b>Dal {data_inizio.strftime('%d/%m/%Y')} al {data_fine.strftime('%d/%m/%Y')}</b>",
            date_style
        ))
        elements.append(Spacer(1, 0.5*cm))
        
        # Informazioni generali
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2E75B5'),
            spaceAfter=10
        )
        
        elements.append(Paragraph("<b>INFORMAZIONI CORSO</b>", info_style))
        
        # Estrai info uniche
        aule_usate = sorted(set(t['aula'] for t in turni if t['aula']))
        formatori_usati = sorted(set(
            f for t in turni 
            for f in [t.get('formatore1'), t.get('formatore2')] 
            if f
        ))
        
        attivita_usate = sorted(set(t['attivita'] for t in turni if t['attivita']))
        
        info_data = [
            ['Durata', f'{giorni_unici} giorni ({len(turni)} turni)'],
            ['Aule', ', '.join(aule_usate) if aule_usate else '-'],
            ['Formatori', ', '.join(formatori_usati) if formatori_usati else '-'],
            ['Attività', ', '.join(attivita_usate) if attivita_usate else '-'],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D9E1F2')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 1*cm))
        
        # Calendario dettagliato
        elements.append(Paragraph("<b>CALENDARIO DETTAGLIATO</b>", info_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Organizza turni per giorno (mattina e pomeriggio separati)
        giorni_dict = defaultdict(lambda: {'mattina': None, 'pomeriggio': None})
        
        for turno in turni_ordinati:
            data_key = turno['data'].date()
            turno_tipo = turno['turno'].lower()
            giorni_dict[data_key][turno_tipo] = turno
        
        # Tabella: colonne per ogni giorno, righe per mattina/pomeriggio
        # MASSIMO 5 GIORNI PER TABELLA
        giorni_ordinati = sorted(giorni_dict.keys())
        
        # Dividi i giorni in gruppi di massimo 5
        for gruppo_idx in range(0, len(giorni_ordinati), 5):
            giorni_gruppo = giorni_ordinati[gruppo_idx:gruppo_idx + 5]
            
            if gruppo_idx > 0:
                # Aggiungi spazio tra le tabelle
                elements.append(Spacer(1, 0.8*cm))
            
            # Header: Date
            header_row = ['Orario'] + [g.strftime('%d/%m\n%a')[:10] for g in giorni_gruppo]
            table_data = [header_row]
            
            # Riga Mattina
            mattina_row = ['Mattina\n' + turno_a_orario('mattina')]
            for giorno in giorni_gruppo:
                turno = giorni_dict[giorno]['mattina']
                if turno:
                    formatori = []
                    if turno.get('formatore1'):
                        formatori.append(turno['formatore1'])
                    if turno.get('formatore2'):
                        formatori.append(turno['formatore2'])
                    
                    # Formato migliorato con etichette
                    attivita = turno.get('attivita', '-')
                    aula = turno.get('aula', '-')
                    
                    cell_text = f"<b>Attività:</b> {attivita}\n"
                    cell_text += f"<b>Aula:</b> {aula}\n"
                    if formatori:
                        cell_text += f"<b>Formatore:</b>\n" + '\n'.join(formatori)
                    else:
                        cell_text += "<b>Formatore:</b> -"
                    
                    # Usa Paragraph per supportare tag HTML
                    cell_style = ParagraphStyle(
                        'CellStyle',
                        parent=styles['Normal'],
                        fontSize=7,
                        alignment=TA_CENTER,
                        leading=9
                    )
                    mattina_row.append(Paragraph(cell_text, cell_style))
                else:
                    mattina_row.append('-')
            
            table_data.append(mattina_row)
            
            # Riga Pomeriggio
            pomeriggio_row = ['Pomeriggio\n' + turno_a_orario('Pomeriggio')]
            for giorno in giorni_gruppo:
                turno = giorni_dict[giorno]['pomeriggio']
                if turno:
                    formatori = []
                    if turno.get('formatore1'):
                        formatori.append(turno['formatore1'])
                    if turno.get('formatore2'):
                        formatori.append(turno['formatore2'])
                    
                    # Formato migliorato con etichette
                    attivita = turno.get('attivita', '-')
                    aula = turno.get('aula', '-')
                    
                    cell_text = f"<b>Attività:</b> {attivita}\n"
                    cell_text += f"<b>Aula:</b> {aula}\n"
                    if formatori:
                        cell_text += f"<b>Formatore:</b>\n" + '\n'.join(formatori)
                    else:
                        cell_text += "<b>Formatore:</b> -"
                    
                    # Usa Paragraph per supportare tag HTML
                    cell_style = ParagraphStyle(
                        'CellStyle',
                        parent=styles['Normal'],
                        fontSize=7,
                        alignment=TA_CENTER,
                        leading=9
                    )
                    pomeriggio_row.append(Paragraph(cell_text, cell_style))
                else:
                    pomeriggio_row.append('-')
            
            table_data.append(pomeriggio_row)
            
            # Calcola larghezza colonne: massimo 5 giorni
            num_giorni = len(giorni_gruppo)
            col_width = 3*cm  # Larghezza fissa per 5 giorni
            
            col_widths = [2.5*cm] + [col_width] * num_giorni
            
            cal_table = Table(table_data, colWidths=col_widths)
            cal_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D9E1F2')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (0, -1), 7),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 1), (-1, -1), 3),
                ('RIGHTPADDING', (0, 1), (-1, -1), 3),
            ]))
            
            elements.append(cal_table)
        
        # Legenda attività (più piccola)
        elements.append(Spacer(1, 0.8*cm))
        
        legend_title_style = ParagraphStyle(
            'LegendTitle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#1F4E78'),
        )
        
        elements.append(Paragraph("<b>LEGENDA ATTIVITÀ:</b>", legend_title_style))
        elements.append(Spacer(1, 0.1*cm))
        
        # Legenda su 2 colonne
        legenda_data = [
            ['AULA', 'Lezione in aula', 'TT', 'Test Tedesco'],
            ['DIGI', 'Attività digitali', 'TI', 'Test Inglese'],
            ['CV', 'Curriculum Vitae', 'UFF', 'Ufficio'],
            ['CS', 'Colloquio selezione', 'COL', 'Colloquio'],
            ['RA', 'Ricerca attiva', 'C', 'Corso']
        ]
        
        legenda_table = Table(legenda_data, colWidths=[1*cm, 3.5*cm, 1*cm, 3.5*cm])
        legenda_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F0F8')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E8F0F8')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(legenda_table)
        
        doc.build(elements)
        print(f"   ✅ Generato corso {nome_percorso}\n")
    
    print(f"✅ Report corsi completati in: {output_dir}/\n")


def genera_report_settimanale(dati, output_dir='stampe_pdf'):
    """
    REPORT 4: Piano Settimanale Completo
    Un PDF per settimana con TUTTE le informazioni (aule, formatori, corsi, fuori aula)
    """
    print("=" * 70)
    print("📅 REPORT SETTIMANALE: PIANO COMPLETO")
    print("=" * 70)
    print()
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Organizza dati per settimana
    settimane = defaultdict(list)
    
    for riga in dati:
        if not riga['data']:
            continue
        
        data = riga['data']
        anno = data.year
        settimana = data.isocalendar()[1]
        chiave = f"{anno}_W{settimana:02d}"
        settimane[chiave].append(riga)
    
    # Genera un PDF per ogni settimana
    for settimana_key in sorted(settimane.keys()):
        turni = settimane[settimana_key]
        if not turni:
            continue
        
        prima_data = turni[0]['data']
        ultima_data = turni[-1]['data']
        anno = prima_data.year
        num_settimana = prima_data.isocalendar()[1]
        
        filename = f"{output_dir}/Piano_Settimanale_W{num_settimana:02d}_{anno}.pdf"
        
        print(f"📄 Settimana {num_settimana} ({prima_data.strftime('%d/%m')} - {ultima_data.strftime('%d/%m/%Y')})")
        
        doc = SimpleDocTemplate(
            filename,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1*cm
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a5490'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        story.append(Paragraph(
            f"PIANO SETTIMANALE - Settimana {num_settimana}/{anno}<br/>"
            f"<font size=10>Dal {prima_data.strftime('%d/%m/%Y')} al {ultima_data.strftime('%d/%m/%Y')}</font>",
            title_style
        ))
        story.append(Spacer(1, 0.5*cm))
        
        table_data = [['Data', 'Turno', 'Percorso 1', 'Percorso 2', 'Percorso 3', 'Percorso 4', 'Fuori Aula']]
        
        for riga in turni:
            data_str = riga['data'].strftime('%d/%m')
            turno = riga['turno']
            
            percorsi_txt = []
            for i in range(4):
                if i < len(riga['percorsi']):
                    nome_perc, perc = riga['percorsi'][i]
                    txt = f"{perc.get('nome', '-')}\n"
                    txt += f"Form: {perc.get('formatore1', '-')}"
                    if perc.get('formatore2'):
                        txt += f", {perc['formatore2']}"
                    txt += f"\nAula: {perc.get('aula', '-')}\n"
                    txt += f"Att: {perc.get('attivita', '-')}"
                    if perc.get('test'):
                        txt += f"\nTest: {perc['test']}"
                    percorsi_txt.append(txt)
                else:
                    percorsi_txt.append('-')
            
            fa_txt = ""
            if riga.get('fuori_aula'):
                fa_txt = "Form: " + ", ".join(riga['fuori_aula'])
            if riga.get('attivita_esterne'):
                if fa_txt:
                    fa_txt += "\n"
                fa_txt += "Att: " + ", ".join(riga['attivita_esterne'])
            if not fa_txt:
                fa_txt = "-"
            
            table_data.append([
                data_str,
                turno,
                percorsi_txt[0],
                percorsi_txt[1],
                percorsi_txt[2],
                percorsi_txt[3],
                fa_txt
            ])
        
        table = Table(table_data, colWidths=[2*cm, 2.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 4*cm])
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1a5490')),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.5*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey)
        story.append(Paragraph("Form = Formatore | Att = Attività | Test = Formatore TEST", footer_style))
        
        doc.build(story)
        print(f"   ✅ Salvato: {filename}")
    
    print(f"\n✅ Generati {len(settimane)} report settimanali")
    print()


def genera_report_formatore_specifico(dati, output_dir='stampe_pdf'):
    """
    REPORT 5: Programma Formatore Specifico con periodo personalizzato
    """
    print("=" * 70)
    print("👤 REPORT FORMATORE SPECIFICO")
    print("=" * 70)
    print()
    
    # Mostra formatori disponibili
    formatori_presenti = set()
    for riga in dati:
        for _, perc in riga['percorsi']:
            if perc.get('formatore1'):
                formatori_presenti.add(perc['formatore1'])
            if perc.get('formatore2'):
                formatori_presenti.add(perc['formatore2'])
        if riga.get('fuori_aula'):
            formatori_presenti.update(riga['fuori_aula'])
    
    formatori_list = sorted(formatori_presenti)
    print("Formatori disponibili:")
    for i, f in enumerate(formatori_list, 1):
        print(f"  {i}. {f}")
    print()
    
    try:
        num = int(input(f"Scegli formatore (1-{len(formatori_list)}): ").strip())
        if num < 1 or num > len(formatori_list):
            print("❌ Numero non valido")
            return
        formatore = formatori_list[num - 1]
    except (ValueError, KeyboardInterrupt, EOFError):
        print("\n❌ Operazione annullata")
        return
    
    print(f"\n📅 Periodo per {formatore}")
    try:
        data_inizio_str = input("Data inizio (gg/mm/aaaa o invio per tutte): ").strip()
        data_fine_str = input("Data fine (gg/mm/aaaa o invio per tutte): ").strip()
        
        data_inizio = None
        data_fine = None
        
        if data_inizio_str:
            data_inizio = datetime.strptime(data_inizio_str, '%d/%m/%Y')
        if data_fine_str:
            data_fine = datetime.strptime(data_fine_str, '%d/%m/%Y')
    except (ValueError, KeyboardInterrupt, EOFError):
        print("\n❌ Formato data non valido o operazione annullata")
        return
    
    # Filtra dati per formatore e periodo
    dati_filtrati = []
    for riga in dati:
        if not riga['data']:
            continue
        
        # Controllo periodo
        if data_inizio and riga['data'] < data_inizio:
            continue
        if data_fine and riga['data'] > data_fine:
            continue
        
        # Controllo se formatore è presente
        presente = False
        for _, perc in riga['percorsi']:
            if perc.get('formatore1') == formatore or perc.get('formatore2') == formatore:
                presente = True
                break
        
        if not presente and riga.get('fuori_aula'):
            if formatore in riga['fuori_aula']:
                presente = True
        
        if presente:
            dati_filtrati.append(riga)
    
    if not dati_filtrati:
        print(f"\n⚠️  Nessun turno trovato per {formatore} nel periodo selezionato")
        return
    
    # Genera PDF
    os.makedirs(output_dir, exist_ok=True)
    
    periodo_str = ""
    if data_inizio and data_fine:
        periodo_str = f"_{data_inizio.strftime('%d%m%Y')}-{data_fine.strftime('%d%m%Y')}"
    elif data_inizio:
        periodo_str = f"_dal_{data_inizio.strftime('%d%m%Y')}"
    elif data_fine:
        periodo_str = f"_al_{data_fine.strftime('%d%m%Y')}"
    
    filename = f"{output_dir}/Programma_Formatore_{formatore}{periodo_str}_2026.pdf"
    
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a5490'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    periodo_text = "Anno 2026"
    if data_inizio and data_fine:
        periodo_text = f"Dal {data_inizio.strftime('%d/%m/%Y')} al {data_fine.strftime('%d/%m/%Y')}"
    elif data_inizio:
        periodo_text = f"Dal {data_inizio.strftime('%d/%m/%Y')}"
    elif data_fine:
        periodo_text = f"Al {data_fine.strftime('%d/%m/%Y')}"
    
    story.append(Paragraph(
        f"PROGRAMMA FORMATORE: {formatore}<br/>"
        f"<font size=12>{periodo_text}</font>",
        title_style
    ))
    story.append(Spacer(1, 1*cm))
    
    # Tabella turni
    table_data = [['Data', 'Turno', 'Percorso', 'Aula', 'Attività', 'Note']]
    
    for riga in dati_filtrati:
        data_str = riga['data'].strftime('%d/%m/%Y')
        turno = riga['turno']
        
        # Trova in quale percorso è il formatore
        for nome_perc, perc in riga['percorsi']:
            if perc.get('formatore1') == formatore or perc.get('formatore2') == formatore:
                ruolo = "Form1" if perc.get('formatore1') == formatore else "Form2"
                altro_form = perc.get('formatore2') if ruolo == "Form1" else perc.get('formatore1')
                note = f"Con: {altro_form}" if altro_form else ""
                
                table_data.append([
                    data_str,
                    turno,
                    perc.get('nome', '-'),
                    perc.get('aula', '-'),
                    perc.get('attivita', '-'),
                    note
                ])
        
        # Fuori aula
        if riga.get('fuori_aula') and formatore in riga['fuori_aula']:
            idx = riga['fuori_aula'].index(formatore)
            att = riga.get('attivita_esterne', [])[idx] if idx < len(riga.get('attivita_esterne', [])) else '-'
            table_data.append([
                data_str,
                turno,
                'FUORI AULA',
                '-',
                att,
                ''
            ])
    
    table = Table(table_data, colWidths=[3*cm, 3*cm, 4*cm, 2.5*cm, 4*cm, 4.5*cm])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1a5490')),
    ]))
    
    story.append(table)
    doc.build(story)
    
    print(f"✅ PDF generato: {filename}")
    print(f"   Turni trovati: {len(dati_filtrati)}")
    print()


def genera_report_corso_specifico(dati, output_dir='stampe_pdf'):
    """
    REPORT 6: Programma Corso Specifico
    """
    print("=" * 70)
    print("📖 REPORT CORSO SPECIFICO")
    print("=" * 70)
    print()
    
    # Raccogli tutti i nomi/numeri di corsi presenti
    corsi_presenti = set()
    for riga in dati:
        for nome_perc, perc in riga['percorsi']:
            if perc.get('nome'):
                corsi_presenti.add(perc['nome'])
    
    corsi_list = sorted(corsi_presenti)
    if not corsi_list:
        print("⚠️  Nessun corso trovato nel file")
        return
    
    print("Corsi disponibili:")
    for i, c in enumerate(corsi_list, 1):
        print(f"  {i}. {c}")
    print()
    
    try:
        num = int(input(f"Scegli corso (1-{len(corsi_list)}): ").strip())
        if num < 1 or num > len(corsi_list):
            print("❌ Numero non valido")
            return
        corso = corsi_list[num - 1]
    except (ValueError, KeyboardInterrupt, EOFError):
        print("\n❌ Operazione annullata")
        return
    
    # Filtra dati per corso
    dati_filtrati = []
    for riga in dati:
        if not riga['data']:
            continue
        
        for nome_perc, perc in riga['percorsi']:
            if perc.get('nome') == corso:
                dati_filtrati.append(riga)
                break
    
    if not dati_filtrati:
        print(f"\n⚠️  Nessun turno trovato per il corso: {corso}")
        return
    
    # Genera PDF
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f"{output_dir}/Programma_Corso_{corso.replace('/', '-')}_2026.pdf"
    
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a5490'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    prima_data = dati_filtrati[0]['data']
    ultima_data = dati_filtrati[-1]['data']
    
    story.append(Paragraph(
        f"PROGRAMMA CORSO: {corso}<br/>"
        f"<font size=12>Dal {prima_data.strftime('%d/%m/%Y')} al {ultima_data.strftime('%d/%m/%Y')}</font>",
        title_style
    ))
    story.append(Spacer(1, 1*cm))
    
    # Tabella turni
    table_data = [['Data', 'Turno', 'Formatori', 'Aula', 'Attività', 'Test']]
    
    for riga in dati_filtrati:
        data_str = riga['data'].strftime('%d/%m/%Y')
        turno = riga['turno']
        
        for nome_perc, perc in riga['percorsi']:
            if perc.get('nome') == corso:
                formatori = perc.get('formatore1', '-')
                if perc.get('formatore2'):
                    formatori += f", {perc['formatore2']}"
                
                table_data.append([
                    data_str,
                    turno,
                    formatori,
                    perc.get('aula', '-'),
                    perc.get('attivita', '-'),
                    perc.get('test', '-')
                ])
                break
    
    table = Table(table_data, colWidths=[3*cm, 3*cm, 5*cm, 3*cm, 5*cm, 2*cm])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1a5490')),
    ]))
    
    story.append(table)
    
    # Totale giorni
    giorni_unici = len(set(r['data'].date() for r in dati_filtrati))
    story.append(Spacer(1, 0.5*cm))
    note_style = ParagraphStyle('Note', parent=styles['Normal'], fontSize=10)
    story.append(Paragraph(f"<b>Totale giorni corso:</b> {giorni_unici}", note_style))
    
    doc.build(story)
    
    print(f"✅ PDF generato: {filename}")
    print(f"   Giorni corso: {giorni_unici}")
    print()


def main():
    print("\n")
    print("╔═══════════════════════════════════════════════════════════════════════╗")
    print("║                                                                       ║")
    print("║           📄 GENERATORE STAMPE PDF - PIANIFICAZIONE 2026              ║")
    print("║                                                                       ║")
    print("╚═══════════════════════════════════════════════════════════════════════╝")
    print("\n")
    
    filename = 'Pianificazione_Corsi_2026.xlsx'
    
    if not os.path.exists(filename):
        print(f"❌ File non trovato: {filename}")
        print("   Assicurati che il file esista nella directory corrente.")
        return
    
    # MENU INTERATTIVO
    print("📋 Quali report vuoi generare?\n")
    print("  1. 📋 Prenotazione Aule (scegli periodo)")
    print("  2. 👥 Programma Formatori (un PDF per ogni formatore)")
    print("  3. 📚 Programma Corsi (un PDF per ogni percorso)")
    print("  4. 📅 Piano Settimanale (un PDF per ogni settimana con TUTTE le info)")
    print("  5. 👤 Programma Formatore SPECIFICO (scegli formatore e periodo)")
    print("  6. 📖 Programma Corso SPECIFICO (scegli corso)")
    print("  7. ✅ TUTTI i report")
    print("  0. ❌ Esci\n")
    
    try:
        scelta = input("Scegli opzione (0-7): ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\n\n❌ Operazione annullata.")
        return
    
    if scelta == '0':
        print("\n❌ Operazione annullata.")
        return
    
    if scelta not in ['1', '2', '3', '4', '5', '6', '7']:
        print(f"\n❌ Scelta non valida: {scelta}")
        return
    
    # Carica dati
    dati = carica_dati_excel(filename)
    
    if not dati:
        print("⚠️  Nessun dato trovato nel file Excel.")
        print("   Compila il file prima di generare i report.")
        return
    
    output_dir = 'stampe_pdf'
    print(f"\n📁 Output directory: {output_dir}/\n")
    
    # Genera i report scelti
    if scelta == '1':
        # Chiedi il periodo per le aule
        print("📅 Seleziona il periodo per la prenotazione aule:\n")
        
        # Trova la prima e ultima data disponibili
        date_disponibili = sorted([r['data'] for r in dati if r['data']])
        if date_disponibili:
            prima_data = date_disponibili[0]
            ultima_data = date_disponibili[-1]
            
            print(f"Date disponibili: {prima_data.strftime('%d/%m/%Y')} - {ultima_data.strftime('%d/%m/%Y')}\n")
            print("Opzioni:")
            print("  1. Tutto l'anno")
            print("  2. Periodo personalizzato")
            print()
            
            try:
                scelta_periodo = input("Scegli opzione (1-2): ").strip()
                
                if scelta_periodo == '1':
                    genera_report_aule(dati, output_dir)
                elif scelta_periodo == '2':
                    print("\nInserisci le date nel formato GG/MM/AAAA")
                    data_inizio_str = input("Data inizio (es. 06/01/2026): ").strip()
                    data_fine_str = input("Data fine (es. 31/01/2026): ").strip()
                    
                    try:
                        data_inizio = datetime.strptime(data_inizio_str, '%d/%m/%Y')
                        data_fine = datetime.strptime(data_fine_str, '%d/%m/%Y')
                        
                        if data_inizio > data_fine:
                            print("❌ La data di inizio deve essere precedente alla data di fine!")
                        else:
                            genera_report_aule(dati, output_dir, data_inizio, data_fine)
                    except ValueError:
                        print("❌ Formato data non valido! Usa GG/MM/AAAA")
                else:
                    print("❌ Scelta non valida")
            except (KeyboardInterrupt, EOFError):
                print("\n❌ Operazione annullata")
        else:
            print("❌ Nessuna data disponibile nei dati")
    
    if scelta == '7':
        genera_report_aule(dati, output_dir)
    
    if scelta == '2' or scelta == '7':
        genera_report_formatori(dati, output_dir)
    
    if scelta == '3' or scelta == '7':
        genera_report_corsi(dati, output_dir)
    
    if scelta == '4' or scelta == '7':
        genera_report_settimanale(dati, output_dir)
    
    if scelta == '5':
        genera_report_formatore_specifico(dati, output_dir)
    
    if scelta == '6':
        genera_report_corso_specifico(dati, output_dir)
    
    # Riepilogo finale
    print("\n")
    print("=" * 70)
    print("✅ GENERAZIONE COMPLETATA!")
    print("=" * 70)
    print()
    print(f"📁 I PDF sono stati salvati in: {output_dir}/")
    print()
    
    if scelta == '1':
        print("📋 Report generato:")
        print("   • Prenotazione aule (periodo selezionato)")
    elif scelta == '2':
        print("📋 Report generati:")
        print("   • Programma_Formatore_[FORMATORE]_2026.pdf (per ogni formatore)")
    elif scelta == '3':
        print("📋 Report generati:")
        print("   • Programma_Corso_[PERCORSO]_2026.pdf (per ogni percorso)")
    elif scelta == '4':
        print("📋 Report generati:")
        print("   • Piano_Settimanale_W[##]_2026.pdf (per ogni settimana)")
    elif scelta == '5':
        print("📋 Report generato:")
        print("   • Programma_Formatore_[FORMATORE]_[PERIODO]_2026.pdf")
    elif scelta == '6':
        print("📋 Report generato:")
        print("   • Programma_Corso_[CORSO]_2026.pdf")
    else:  # scelta == '7'
        print("📋 Report generati:")
        print("   1. Prenotazione_Aule_2026.pdf (tutte le aule)")
        print("   2. Programma_Formatore_[FORMATORE]_2026.pdf (per ogni formatore)")
        print("   3. Programma_Corso_[PERCORSO]_2026.pdf (per ogni percorso)")
        print("   4. Piano_Settimanale_W[##]_2026.pdf (per ogni settimana)")
    
    print()
    print("🎉 Sistema pronto!")
    print("=" * 70)
    print()


if __name__ == '__main__':
    main()
