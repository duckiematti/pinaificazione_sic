#!/usr/bin/env python3
"""
PIANIFICAZIONE CORSI 2026 - SISTEMA INTELLIGENTE
=================================================

STRATEGIA:
1. DataValidation per liste base (formatori, aule)
2. Conditional Formatting per evidenziare DUPLICATI in rosso
3. Sezioni AULE separate per filtraggio attività
4. Celle di supporto per validazione dinamica

"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, timedelta
import calendar

# CONFIGURAZIONE
FORMATORI = ['CL', 'MC', 'LD', 'EP', 'IP', 'FB', 'GZ', 'DC']
FORMATORI_TEST = ['URS', 'NIC', 'MIT', 'MON', 'WER']

# ATTIVITÀ ESTERNE (per fuori aula)
ATTIVITA_ESTERNE = ['RIUNIONE', 'FORMAZIONE', 'CONSULENZA', 'AUDIT', 'ALTRO']

# AULE con ATTIVITÀ COMPATIBILI - LA CHIAVE DEL SISTEMA
AULE_ATTIVITA = {
    '103': ['AULA', 'DIGI', 'CV', 'CS', 'RA', 'TT', 'TI'],
    '103a': ['AULA', 'DIGI', 'CV', 'CS', 'RA', 'TT', 'TI'],
    '108': ['AULA', 'CV', 'CS', 'RA'],
    '110': ['AULA', 'CV', 'CS', 'RA'],
    'UFF': ['UFF', 'COL', 'C']
}

# FESTIVITÀ 2026
FESTIVITA = [
    datetime(2026, 1, 1),   # Capodanno
    datetime(2026, 1, 6),   # Epifania
    datetime(2026, 4, 6),   # Pasquetta
    datetime(2026, 4, 25),  # Liberazione
    datetime(2026, 5, 1),   # Lavoro
    datetime(2026, 6, 2),   # Repubblica
    datetime(2026, 8, 15),  # Ferragosto
    datetime(2026, 11, 1),  # Ognissanti
    datetime(2026, 12, 8),  # Immacolata
    datetime(2026, 12, 25), # Natale
    datetime(2026, 12, 26), # S. Stefano
    # Ferie aziendali
    datetime(2026, 8, 10), datetime(2026, 8, 11), datetime(2026, 8, 12),
    datetime(2026, 8, 13), datetime(2026, 8, 14),
]

# COLORI
COLOR_HEADER = 'D9E1F2'
COLOR_MORNING = 'E2EFDA'
COLOR_AFTERNOON = 'FFF2CC'
COLOR_MONTH = 'B4C7E7'
COLOR_DUPLICATE = 'FF0000'  # Rosso per duplicati
COLOR_WARNING = 'FFC7CE'    # Rosa per warning

# 40 colori per percorsi
COLORI_PERCORSI = [
    'B4C7E7', 'F8CBAD', 'C5E0B4', 'FFE699', 'B4C7E7', 'D5A6BD', 
    'A9D08E', 'F4B084', 'BDD7EE', 'F8CBAD', 'C6E0B4', 'FFD966',
    '9DC3E6', 'F4B183', 'A8D08D', 'FFEB9C', '8FAADC', 'E2A293',
    '9BBB59', 'FFD556', '7FA7D0', 'D99694', '92D050', 'FFC000',
    '6FA8DC', 'CC8899', '76A35D', 'F9CB9C', '5B9BD5', 'B38EAC',
    '70AD47', 'ED7D31', '4A7EBB', 'A87B9C', '548235', 'C65911',
    '385D8A', '9B6B81', '375623', 'A04D00'
]

def is_weekend(date):
    return date.weekday() >= 5

def is_holiday(date):
    return date in FESTIVITA

def create_assumptions_sheet(wb):
    """Foglio Assumptions - Festività e Orari Lezioni"""
    ws = wb.create_sheet('Assumptions', 0)
    
    # SEZIONE ORARI LEZIONI
    ws['A1'] = 'IMPOSTAZIONI ORARI LEZIONI'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E75B5', end_color='2E75B5', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:C1')
    
    # Intestazioni
    ws['A3'] = 'Turno'
    ws['B3'] = 'Orario Inizio'
    ws['C3'] = 'Orario Fine'
    for col in ['A3', 'B3', 'C3']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws[col].alignment = Alignment(horizontal='center')
    
    # Dati orari
    ws['A4'] = 'Mattina'
    ws['B4'] = '09:00'
    ws['C4'] = '13:00'
    
    ws['A5'] = 'Pomeriggio'
    ws['B5'] = '14:00'
    ws['C5'] = '18:00'
    
    # Formattazione celle orari
    for row in [4, 5]:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
    
    # Dimensioni colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # SEZIONE FESTIVITÀ
    ws['A8'] = 'GIORNI DA ESCLUDERE (FESTIVITÀ E FERIE)'
    ws['A8'].font = Font(bold=True, size=12)
    ws['A8'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    for idx, festivita in enumerate(sorted(FESTIVITA), start=10):
        ws[f'A{idx}'] = festivita
        ws[f'A{idx}'].number_format = 'DD/MM/YYYY'
    
    ws.column_dimensions['A'].width = 30

def create_formatori_sheet(wb):
    """Foglio FORMATORI - Con formule per conteggio automatico dal foglio 2026"""
    ws = wb.create_sheet('FORMATORI')
    
    # Headers
    headers = ['FORMATORI', '%', 'n.giorni\nprevisti', 'Settimana\nnon lavoro', 
               'Festività e ferie', 'n.giorni\ndisponibili', 'n.giorni\nsvolti', 
               'n.giorni\nrimanenti', 'FORMATORI TEST']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col_idx == 7:  # n.giorni svolti
            cell.fill = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
    
    # Dati formatori principali
    formatori_data = [
        ('CL', 0.7, 155, '', 0),
        ('MC', 0.5, 111, 'Mercoledì Mattina; Mercoledì pomeriggio', 0),
        ('LD', 0.8, 177, '', 0),
        ('EP', 0.9, 199, '', 0),
        ('IP', 0.9, 199, '', 0),
        ('FB', 0.8, 177, '', 0),
        ('GZ', 0.8, 177, '', 0),
        ('DC', 0.4, 88, 'Mercoledì Mattina; Mercoledì pomeriggio', 0),
    ]
    
    for idx, (nome, perc, giorni, non_lavoro, ferie) in enumerate(formatori_data, start=2):
        ws[f'A{idx}'] = nome
        ws[f'B{idx}'] = perc
        ws[f'B{idx}'].number_format = '0%'
        ws[f'C{idx}'] = giorni
        ws[f'D{idx}'] = non_lavoro
        ws[f'E{idx}'] = ferie
        ws[f'F{idx}'] = f'=C{idx}-E{idx}'
        
        # FORMULA CONTA GIORNI: conta quante volte il formatore appare nel foglio 2026
        # Solo colonne formatori: D,E,J,K,P,Q,V,W,AB,AD,AF,AH,AJ (28,30,32,34,36)
        ws[f'G{idx}'] = (
            f'=SUM('
            f'COUNTIF(\'2026\'!D:D,A{idx}),COUNTIF(\'2026\'!E:E,A{idx}),'
            f'COUNTIF(\'2026\'!J:J,A{idx}),COUNTIF(\'2026\'!K:K,A{idx}),'
            f'COUNTIF(\'2026\'!P:P,A{idx}),COUNTIF(\'2026\'!Q:Q,A{idx}),'
            f'COUNTIF(\'2026\'!V:V,A{idx}),COUNTIF(\'2026\'!W:W,A{idx}),'
            f'COUNTIF(\'2026\'!AB:AB,A{idx}),COUNTIF(\'2026\'!AD:AD,A{idx}),'
            f'COUNTIF(\'2026\'!AF:AF,A{idx}),COUNTIF(\'2026\'!AH:AH,A{idx}),COUNTIF(\'2026\'!AJ:AJ,A{idx})'
            f')'
        )
        ws[f'G{idx}'].fill = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
        
        ws[f'H{idx}'] = f'=F{idx}-G{idx}'
    
    # Formatori TEST (ora integrati sopra, questa sezione mantiene la colonna I per referenza)
    for idx, test_nome in enumerate(FORMATORI_TEST, start=2):
        ws[f'I{idx}'] = test_nome
    
    # Liste nascoste per validazione (riga 50+)
    for idx, nome in enumerate(FORMATORI, start=51):
        ws[f'A{idx}'] = nome
    
    for idx, test_nome in enumerate(FORMATORI_TEST, start=51):
        ws[f'I{idx}'] = test_nome
    
    # Nascondi righe helper
    for row in range(50, 70):
        ws.row_dimensions[row].hidden = True
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['G'].width = 15

def create_controllo_aule_sheet(wb):
    """Foglio CONTROLLO_AULE - Con sezioni per ogni aula"""
    ws = wb.create_sheet('CONTROLLO_AULE')
    
    # STRATEGIA: Creare range nominati per ogni aula
    # Esempio: 103_ATTIVITA contiene solo le attività compatibili con aula 103
    
    ws['A1'] = 'AULA'
    ws['B1'] = 'DESCRIZIONE'
    ws['C1'] = 'ATTIVITÀ COMPATIBILI'
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = Font(bold=True)
    
    current_row = 2
    
    # Per ogni aula, lista le attività compatibili
    for aula, attivita in AULE_ATTIVITA.items():
        ws[f'A{current_row}'] = aula
        ws[f'B{current_row}'] = f'Aula {aula}'
        
        # Metti tutte le attività in colonna C
        for idx, att in enumerate(attivita):
            ws[f'C{current_row + idx}'] = att
        
        current_row += len(attivita) + 1  # Spazio tra aule
    
    # Liste per validazione (riga 50+)
    aule_list = list(AULE_ATTIVITA.keys())
    for idx, aula in enumerate(aule_list, start=50):
        ws[f'A{idx}'] = aula
    
    # Tutte le attività (per lista completa)
    all_att = sorted(set(att for atts in AULE_ATTIVITA.values() for att in atts))
    for idx, att in enumerate(all_att, start=50):
        ws[f'C{idx}'] = att
    
    # Attività esterne (colonna D)
    for idx, att_est in enumerate(ATTIVITA_ESTERNE, start=50):
        ws[f'D{idx}'] = att_est
    
    # SEZIONI SEPARATE PER OGNI AULA (per validazione dinamica)
    start_col = 5  # Colonna E
    for aula, attivita in AULE_ATTIVITA.items():
        col_letter = get_column_letter(start_col)
        ws[f'{col_letter}1'] = f'{aula}_ATT'
        ws[f'{col_letter}1'].font = Font(bold=True)
        
        for idx, att in enumerate(attivita, start=50):
            ws[f'{col_letter}{idx}'] = att
        
        start_col += 1
    
    # Nascondi righe helper
    for row in range(50, 70):
        ws.row_dimensions[row].hidden = True
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

def create_att_esterne_sheet(wb):
    """Foglio ATT. ESTERNE - Personalizzabile dall'utente"""
    ws = wb.create_sheet('ATT. ESTERNE')
    
    # TITOLO
    ws['A1'] = 'CONFIGURAZIONE ATTIVITÀ ESTERNE'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E75B5', end_color='2E75B5', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:C1')
    
    # INTESTAZIONI
    ws['A3'] = 'Codice'
    ws['B3'] = 'Descrizione'
    ws['C3'] = 'Note'
    for col in ['A3', 'B3', 'C3']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws[col].alignment = Alignment(horizontal='center')
    
    # DATI PREDEFINITI (Modificabili dall'utente)
    attivita = [
        ('Amm', 'Amministrazione', 'Attività amministrative'),
        ('IA', 'Intelligenza Artificiale', 'Corsi IA'),
        ('AI', 'Corso AI', 'Corso specifico AI'),
        ('Dig', 'Digitale', 'Attività digitali'),
        ('P1', 'Progetto 1', 'Modificabile'),
        ('P2', 'Progetto 2', 'Modificabile'),
        ('P3', 'Progetto 3', 'Modificabile'),
        ('P4', 'Progetto 4', 'Modificabile'),
        ('P5', 'Progetto 5', 'Modificabile'),
    ]
    
    for idx, (sigla, desc, note) in enumerate(attivita, start=4):
        ws[f'A{idx}'] = sigla
        ws[f'B{idx}'] = desc
        ws[f'C{idx}'] = note
        
        # Proteggi solo la colonna A (codice)
        ws[f'A{idx}'].font = Font(bold=True)
        
        # Evidenzia i progetti modificabili
        if sigla.startswith('P'):
            ws[f'B{idx}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            ws[f'C{idx}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Range nascosto per validazione (riga 50+)
    for idx, (sigla, desc, note) in enumerate(attivita, start=50):
        ws[f'B{idx}'] = sigla
    
    # Nascondi righe helper
    for row in range(50, 70):
        ws.row_dimensions[row].hidden = True
    
    # Istruzioni
    ws['A15'] = '💡 ISTRUZIONI:'
    ws['A15'].font = Font(bold=True, size=11)
    ws['A16'] = '1. Modifica la colonna "Descrizione" per i progetti P1-P5'
    ws['A17'] = '2. Esempio: P1 = "Progetto Sostenibilità"'
    ws['A18'] = '3. Le modifiche appariranno automaticamente nei PDF dei formatori'
    
    for row in [16, 17, 18]:
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

def create_aule_sheet(wb):
    """Foglio AULE_1"""
    ws = wb.create_sheet('AULE_1')
    
    ws['A1'] = 'AULE DISPONIBILI'
    ws['A1'].font = Font(bold=True)
    
    for idx, aula in enumerate(AULE_ATTIVITA.keys(), start=2):
        ws[f'A{idx}'] = aula
    
    ws.column_dimensions['A'].width = 15

def add_validations_smart(ws, row):
    """
    Aggiunge validazioni INTELLIGENTI con:
    1. DataValidation per liste base
    2. Formule per celle helper
    3. Setup per conditional formatting
    """
    
    # ===== PERCORSO 1 (C-H) =====
    # Formatore 1 (D)
    dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
    dv.error = 'Seleziona un formatore valido'
    dv.errorTitle = 'Formatore non valido'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=4))
    
    # Formatore 2 (E)
    dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
    dv.error = 'Seleziona un formatore valido'
    dv.errorTitle = 'Formatore non valido'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=5))
    
    # Aula (F)
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$A$50:$A$54', allow_blank=True)
    dv.error = 'Seleziona un\'aula valida'
    dv.errorTitle = 'Aula non valida'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=6))
    
    # Attività (G) - CON VALIDAZIONE MIGLIORATA
    # Mostra tutte le attività ma con prompt specifico per compatibilità
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$C$50:$C$65', allow_blank=True)
    dv.error = '⚠️ VERIFICA COMPATIBILITÀ!\n\n103/103a: tutte le attività\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.errorTitle = 'Attività - Verifica Aula'
    dv.prompt = '⚠️ ATTENZIONE:\nSCEGLI PRIMA L\'AULA, POI VERIFICA:\n\n📌 103/103a: TUTTE le attività\n📌 108/110: solo AULA,CV,CS,RA\n📌 UFF: solo UFF,COL,C'
    dv.promptTitle = '🏫 Compatibilità Aula-Attività'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=7))
    
    # Test (H)
    dv = DataValidation(type="list", formula1='=FORMATORI!$I$51:$I$55', allow_blank=True)
    dv.error = 'Solo formatori TEST (per TT/TI)'
    dv.errorTitle = 'Test'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=8))
    
    # ===== PERCORSO 2 (I-N) =====
    for col in [10, 11]:  # Formatori
        dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=col))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$A$50:$A$54', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=12))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$C$50:$C$65', allow_blank=True)
    dv.error = '⚠️ VERIFICA COMPATIBILITÀ!\n103/103a: tutte\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.errorTitle = 'Attività - Verifica Aula'
    dv.prompt = '⚠️ SCEGLI PRIMA L\'AULA!\n\n103/103a: TUTTE\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.promptTitle = '🏫 Compatibilità'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=13))
    
    # Test percorso 2 (N/14)
    dv = DataValidation(type="list", formula1='=FORMATORI!$I$51:$I$55', allow_blank=True)
    dv.error = 'Solo formatori TEST (per TT/TI)'
    dv.errorTitle = 'Test'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=14))
    
    # ===== PERCORSO 3 (O-T) =====
    for col in [16, 17]:
        dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=col))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$A$50:$A$54', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=18))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$C$50:$C$65', allow_blank=True)
    dv.error = '⚠️ VERIFICA COMPATIBILITÀ!\n103/103a: tutte\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.errorTitle = 'Attività - Verifica Aula'
    dv.prompt = '⚠️ SCEGLI PRIMA L\'AULA!\n\n103/103a: TUTTE\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.promptTitle = '🏫 Compatibilità'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=19))
    
    # Test percorso 3 (T/20)
    dv = DataValidation(type="list", formula1='=FORMATORI!$I$51:$I$55', allow_blank=True)
    dv.error = 'Solo formatori TEST (per TT/TI)'
    dv.errorTitle = 'Test'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=20))
    # ===== PERCORSO 4 (U-Z) =====
    for col in [22, 23]:
        dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=col))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$A$50:$A$54', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=24))
    
    dv = DataValidation(type="list", formula1='=CONTROLLO_AULE!$C$50:$C$65', allow_blank=True)
    dv.error = '⚠️ VERIFICA COMPATIBILITÀ!\n103/103a: tutte\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.errorTitle = 'Attività - Verifica Aula'
    dv.prompt = '⚠️ SCEGLI PRIMA L\'AULA!\n\n103/103a: TUTTE\n108/110: AULA,CV,CS,RA\nUFF: UFF,COL,C'
    dv.promptTitle = '🏫 Compatibilità'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=25))
    
    # Test percorso 4 (Z/26)
    dv = DataValidation(type="list", formula1='=FORMATORI!$I$51:$I$55', allow_blank=True)
    dv.error = 'Solo formatori TEST (per TT/TI)'
    dv.errorTitle = 'Test'
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=26))
    
    # ===== FUORI AULA (AB-AK) =====
    # Alternanza: Form(28,30,32,34,36) + Att.Est(29,31,33,35,37) - colonne AB-AK
    
    # Formatori: colonne 28, 30, 32, 34, 36 (AB, AD, AF, AH, AJ)
    for col in [28, 30, 32, 34, 36]:
        dv = DataValidation(type="list", formula1='=FORMATORI!$A$51:$A$58', allow_blank=True)
        dv.error = 'Seleziona un formatore valido'
        dv.errorTitle = 'Formatore non valido'
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=col))
    
    # Attività esterne: colonne 29, 31, 33, 35, 37 (AC, AE, AG, AI, AK)
    for col in [29, 31, 33, 35, 37]:
        dv = DataValidation(type="list", formula1="='ATT. ESTERNE'!$B$50:$B$58", allow_blank=True)
        dv.error = 'Seleziona un\'attività esterna valida'
        dv.errorTitle = 'Attività Esterna'
        dv.prompt = 'Attività svolte fuori dalle aule BCC'
        dv.promptTitle = 'Attività Esterne'
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=col))

def add_conditional_formatting(ws, start_row, end_row):
    """
    Aggiunge Conditional Formatting per evidenziare DUPLICATI
    
    LOGICA:
    - Confronta ogni cella formatore con altre celle formatore nella stessa riga
    - Se trova duplicato → sfondo ROSSO
    - Confronta aule per duplicati → sfondo ROSSO
    """
    
    # FORMATORI: colonne D,E,J,K,P,Q,V,W + Fuori aula: AB,AD,AF,AH,AJ (28,30,32,34,36)
    formatori_cols = [4, 5, 10, 11, 16, 17, 22, 23, 28, 30, 32, 34, 36]
    
    for row in range(start_row, end_row + 1):
        for col in formatori_cols:
            cell_ref = f'{get_column_letter(col)}{row}'
            
            # Formula per controllare se lo stesso valore appare in altre celle formatore
            other_cols = [c for c in formatori_cols if c != col]
            formula_parts = [f'{get_column_letter(c)}{row}={cell_ref}' for c in other_cols]
            formula = f'=AND(LEN({cell_ref})>0, OR({",".join(formula_parts)}))'
            
            rule = FormulaRule(formula=[formula], fill=PatternFill(start_color=COLOR_DUPLICATE, end_color=COLOR_DUPLICATE, fill_type='solid'))
            ws.conditional_formatting.add(cell_ref, rule)
    
    # AULE: colonne F, L, R, X (6, 12, 18, 24)
    aule_cols = [6, 12, 18, 24]
    
    for row in range(start_row, end_row + 1):
        for col in aule_cols:
            cell_ref = f'{get_column_letter(col)}{row}'
            
            other_cols = [c for c in aule_cols if c != col]
            formula_parts = [f'{get_column_letter(c)}{row}={cell_ref}' for c in other_cols]
            formula = f'=AND(LEN({cell_ref})>0, OR({",".join(formula_parts)}))'
            
            rule = FormulaRule(formula=[formula], fill=PatternFill(start_color=COLOR_DUPLICATE, end_color=COLOR_DUPLICATE, fill_type='solid'))
            ws.conditional_formatting.add(cell_ref, rule)
    
    # INCOMPATIBILITÀ AULA-ATTIVITÀ: evidenzia in ROSA
    # Percorso 1: Aula F (6), Attività G (7)
    # Percorso 2: Aula L (12), Attività M (13)
    # Percorso 3: Aula R (18), Attività S (19)
    # Percorso 4: Aula X (24), Attività Y (25)
    
    incompatibility_rules = [
        # Formato: (col_aula, col_attivita)
        (6, 7),    # Percorso 1
        (12, 13),  # Percorso 2
        (18, 19),  # Percorso 3
        (24, 25),  # Percorso 4
    ]
    
    for row in range(start_row, end_row + 1):
        for col_aula, col_attivita in incompatibility_rules:
            cell_aula = f'{get_column_letter(col_aula)}{row}'
            cell_att = f'{get_column_letter(col_attivita)}{row}'
            
            # Formula complessa per verificare compatibilità
            # 108/110: solo AULA,CV,CS,RA
            # UFF: solo UFF,COL,C
            # 103/103a: tutte (nessuna incompatibilità)
            formula = (
                f'=AND('
                f'LEN({cell_att})>0, '
                f'OR('
                # 108 con attività non compatibili
                f'AND({cell_aula}="108", NOT(OR({cell_att}="AULA",{cell_att}="CV",{cell_att}="CS",{cell_att}="RA"))),'
                # 110 con attività non compatibili
                f'AND({cell_aula}="110", NOT(OR({cell_att}="AULA",{cell_att}="CV",{cell_att}="CS",{cell_att}="RA"))),'
                # UFF con attività non compatibili
                f'AND({cell_aula}="UFF", NOT(OR({cell_att}="UFF",{cell_att}="COL",{cell_att}="C")))'
                f')'
                f')'
            )
            
            rule = FormulaRule(formula=[formula], fill=PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid'))
            ws.conditional_formatting.add(cell_att, rule)

def create_main_schedule_sheet(wb):
    """Foglio 2026 - SISTEMA INTELLIGENTE"""
    ws = wb.create_sheet('2026', 1)
    
    # Titolo
    ws['A1'] = 'ML5-05 Piano di dettaglio BCC'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['V1'] = f'Aggiornato il {datetime.now().strftime("%d.%m.%Y")}'
    ws['V1'].font = Font(italic=True)
    
    # Istruzioni importanti
    ws['A2'] = '⚠️ IMPORTANTE: Le celle ROSSE indicano DUPLICATI (formatori/aule ripetuti nello stesso turno)'
    ws['A2'].font = Font(bold=True, color='FF0000', size=10)
    ws.merge_cells('A2:Z2')
    
    current_row = 4
    
    # Genera giorni lavorativi
    all_working_days = []
    for month in range(1, 13):
        num_days = calendar.monthrange(2026, month)[1]
        for day in range(1, num_days + 1):
            current_date = datetime(2026, month, day)
            if not is_weekend(current_date) and not is_holiday(current_date):
                all_working_days.append(current_date)
    
    print(f"Giorni lavorativi totali: {len(all_working_days)}")
    
    current_month = None
    first_data_row = None
    
    for date_idx, current_date in enumerate(all_working_days):
        if current_date.month != current_month:
            current_month = current_date.month
            month_name = calendar.month_name[current_month].upper()
            
            # Header mese
            ws.cell(row=current_row, column=1, value=month_name).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color=COLOR_MONTH, end_color=COLOR_MONTH, fill_type='solid')
            ws.merge_cells(f'A{current_row}:B{current_row}')
            
            ws.cell(row=current_row, column=3, value='BCC')
            ws.cell(row=current_row, column=28, value='Fuori aula').font = Font(bold=True)
            
            current_row += 1
            
            # Headers colonne
            headers = [
                (1, '#REF!+1'), (2, 'Turno'),
                # Percorso 1
                (3, 'percorso'), (4, 'Formatore 1'), (5, 'Formatore 2'), 
                (6, 'Aula'), (7, 'Attività'), (8, 'Test'),
                # Percorso 2
                (9, 'percorso'), (10, 'Formatore 1'), (11, 'Formatore 2'), 
                (12, 'Aula'), (13, 'Attività'), (14, 'Test'),
                # Percorso 3
                (15, 'percorso'), (16, 'Formatore 1'), (17, 'Formatore 2'), 
                (18, 'Aula'), (19, 'Attività'), (20, 'Test'),
                # Percorso 4
                (21, 'percorso'), (22, 'Formatore 1'), (23, 'Formatore 2'), 
                (24, 'Aula'), (25, 'Attività'), (26, 'Test'),
                # Fine corso + Fuori aula
                (27, 'Fine corso'),
                # Fuori aula: 5 coppie formatore-attività
                (28, 'Form.1'), (29, 'Att.Est.1'), (30, 'Form.2'), (31, 'Att.Est.2'), 
                (32, 'Form.3'), (33, 'Att.Est.3'), (34, 'Form.4'), (35, 'Att.Est.4'),
                (36, 'Form.5'), (37, 'Att.Est.5')
            ]
            
            for col, header_text in headers:
                cell = ws.cell(row=current_row, column=col, value=header_text)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            current_row += 1
        
        if first_data_row is None:
            first_data_row = current_row
        
        # Riga Mattina
        ws.cell(row=current_row, column=1, value=current_date).number_format = 'DD/MM/YYYY'
        ws.cell(row=current_row, column=2, value='mattina')
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=COLOR_MORNING, end_color=COLOR_MORNING, fill_type='solid')
        
        add_validations_smart(ws, current_row)
        current_row += 1
        
        # Riga Pomeriggio
        ws.cell(row=current_row, column=2, value='Pomeriggio')
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color=COLOR_AFTERNOON, end_color=COLOR_AFTERNOON, fill_type='solid')
        
        add_validations_smart(ws, current_row)
        current_row += 1
    
    last_data_row = current_row - 1
    
    # APPLICA CONDITIONAL FORMATTING per evidenziare DUPLICATI
    print(f"Applicazione Conditional Formatting righe {first_data_row}-{last_data_row}...")
    add_conditional_formatting(ws, first_data_row, last_data_row)
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    for col_idx in range(3, 38):  # Fino a colonna 37 (AK)
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    
    print("✅ Conditional Formatting applicato - Duplicati verranno evidenziati in ROSSO")

def main():
    print("=" * 70)
    print("🎓 PIANIFICAZIONE CORSI 2026 - SISTEMA INTELLIGENTE")
    print("=" * 70)
    print()
    
    wb = Workbook()
    wb.remove(wb.active)
    
    print("📋 Creazione fogli...")
    create_assumptions_sheet(wb)
    create_main_schedule_sheet(wb)
    create_aule_sheet(wb)
    create_formatori_sheet(wb)
    create_controllo_aule_sheet(wb)
    create_att_esterne_sheet(wb)
    
    filename = 'Pianificazione_Corsi_2026.xlsx'
    print(f"\n💾 Salvataggio: {filename}...")
    wb.save(filename)
    
    print()
    print("=" * 70)
    print("✅ FILE CREATO CON SISTEMA INTELLIGENTE!")
    print("=" * 70)
    print()
    print("🎯 FUNZIONALITÀ IMPLEMENTATE:")
    print("  ✅ Menu a tendina per formatori, aule, attività")
    print("  ✅ CONDITIONAL FORMATTING: Duplicati evidenziati in ROSSO")
    print("  ✅ Prompt per ricordare compatibilità aula-attività")
    print("  ✅ Sezioni separate per ogni aula (per future implementazioni)")
    print()
    print("📁 File:", filename)
    print("=" * 70)

if __name__ == '__main__':
    main()
